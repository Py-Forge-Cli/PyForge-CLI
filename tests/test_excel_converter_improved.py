"""Improved tests for Excel converter using generated test data."""

import tempfile
from pathlib import Path
from unittest.mock import patch

import pytest

from pyforge_cli.converters.excel_converter import ExcelConverter


class TestExcelConverterImproved:
    """Improved test suite for Excel converter."""

    @pytest.fixture
    def converter(self):
        """Create an Excel converter instance."""
        return ExcelConverter()

    @pytest.fixture
    def temp_dir(self):
        """Create a temporary directory."""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield Path(tmpdir)

    @pytest.fixture
    def test_data_dir(self):
        """Get the test data directory."""
        return Path(__file__).parent / 'data' / 'excel'

    def test_supported_formats(self, converter):
        """Test supported format detection."""
        # ExcelConverter should have these attributes set
        assert hasattr(converter, 'supported_inputs')
        assert hasattr(converter, 'supported_outputs')

        # Check supported formats if they exist
        if converter.supported_inputs:
            assert '.xlsx' in converter.supported_inputs or '.xls' in converter.supported_inputs

    def test_validate_basic_excel(self, converter, test_data_dir):
        """Test validation of basic Excel file."""
        excel_file = test_data_dir / 'basic.xlsx'
        if excel_file.exists():
            assert converter.validate_input(excel_file) is True

    def test_validate_multi_sheet_excel(self, converter, test_data_dir):
        """Test validation of multi-sheet Excel file."""
        excel_file = test_data_dir / 'multi_sheet.xlsx'
        if excel_file.exists():
            assert converter.validate_input(excel_file) is True

    def test_validate_invalid_file(self, converter, temp_dir):
        """Test validation of non-Excel file."""
        invalid_file = temp_dir / 'not_excel.txt'
        invalid_file.write_text('This is not an Excel file')
        assert converter.validate_input(invalid_file) is False

    def test_validate_nonexistent_file(self, converter):
        """Test validation of non-existent file."""
        assert converter.validate_input(Path('/nonexistent/file.xlsx')) is False

    def test_get_metadata_basic_file(self, converter, test_data_dir):
        """Test metadata extraction from basic Excel file."""
        excel_file = test_data_dir / 'basic.xlsx'
        if excel_file.exists():
            metadata = converter.get_metadata(excel_file)
            assert metadata is not None
            assert isinstance(metadata, dict)
            assert 'file_name' in metadata
            assert 'file_format' in metadata
            assert 'sheet_count' in metadata
            assert 'sheet_names' in metadata
        else:
            pytest.skip("Test data file not found")

    def test_get_metadata_multi_sheet_file(self, converter, test_data_dir):
        """Test metadata extraction from multi-sheet Excel file."""
        excel_file = test_data_dir / 'multi_sheet.xlsx'
        if excel_file.exists():
            metadata = converter.get_metadata(excel_file)
            assert metadata is not None
            assert isinstance(metadata, dict)
            assert metadata['sheet_count'] > 0
            assert len(metadata['sheet_names']) > 0
            assert 'sheet_details' in metadata
        else:
            pytest.skip("Test data file not found")

    def test_get_metadata_empty_sheets_file(self, converter, test_data_dir):
        """Test metadata extraction from Excel with empty sheets."""
        excel_file = test_data_dir / 'empty_sheets.xlsx'
        if excel_file.exists():
            metadata = converter.get_metadata(excel_file)
            assert metadata is not None
            assert isinstance(metadata, dict)
            assert 'sheet_details' in metadata
            # Empty sheets should still be counted
            assert metadata['sheet_count'] > 0
        else:
            pytest.skip("Test data file not found")

    def test_convert_basic_excel(self, converter, test_data_dir, temp_dir):
        """Test converting basic Excel file."""
        excel_file = test_data_dir / 'basic.xlsx'
        output_file = temp_dir / 'output.parquet'

        if excel_file.exists():
            result = converter.convert(excel_file, output_file)
            assert result is True
            assert output_file.exists()

    def test_convert_mixed_types(self, converter, test_data_dir, temp_dir):
        """Test converting Excel with mixed data types."""
        excel_file = test_data_dir / 'mixed_types.xlsx'
        output_file = temp_dir / 'output.parquet'

        if excel_file.exists():
            result = converter.convert(excel_file, output_file)
            assert result is True
            assert output_file.exists()

    def test_convert_special_characters(self, converter, test_data_dir, temp_dir):
        """Test converting Excel with special characters."""
        excel_file = test_data_dir / 'special_characters.xlsx'
        output_file = temp_dir / 'output.parquet'

        if excel_file.exists():
            result = converter.convert(excel_file, output_file)
            assert result is True
            assert output_file.exists()

    def test_convert_null_heavy(self, converter, test_data_dir, temp_dir):
        """Test converting Excel with many null values."""
        excel_file = test_data_dir / 'null_heavy.xlsx'
        output_file = temp_dir / 'output.parquet'

        if excel_file.exists():
            result = converter.convert(excel_file, output_file)
            assert result is True
            assert output_file.exists()

    def test_convert_large_dataset(self, converter, test_data_dir, temp_dir):
        """Test converting large Excel file."""
        excel_file = test_data_dir / 'large_dataset.xlsx'
        output_file = temp_dir / 'output.parquet'

        if excel_file.exists():
            result = converter.convert(excel_file, output_file)
            assert result is True
            assert output_file.exists()

    def test_convert_multi_sheet_separate(self, converter, test_data_dir, temp_dir):
        """Test converting multi-sheet Excel to separate files."""
        excel_file = test_data_dir / 'multi_sheet.xlsx'
        output_dir = temp_dir / 'output'

        if excel_file.exists():
            result = converter.convert(excel_file, output_dir, separate=True)
            assert result is True
            assert output_dir.exists()

            # Check that separate files were created
            expected_files = [
                'multi_sheet_Employees.parquet',
                'multi_sheet_Departments.parquet',
                'multi_sheet_Projects.parquet'
            ]
            for filename in expected_files:
                assert (output_dir / filename).exists()

    def test_convert_multi_sheet_combine(self, converter, test_data_dir, temp_dir):
        """Test combining sheets with matching columns."""
        # Create a test Excel with sheets having same columns
        excel_file = test_data_dir / 'pivot_like.xlsx'
        output_file = temp_dir / 'combined.parquet'

        if excel_file.exists():
            # This file has consistent structure across rows
            result = converter.convert(excel_file, output_file, combine=True)
            assert result is True
            assert output_file.exists()

    def test_convert_with_compression(self, converter, test_data_dir, temp_dir):
        """Test converting with different compression options."""
        excel_file = test_data_dir / 'basic.xlsx'

        if excel_file.exists():
            # Test gzip compression
            output_gzip = temp_dir / 'output_gzip.parquet'
            result = converter.convert(excel_file, output_gzip, compression='gzip')
            assert result is True
            assert output_gzip.exists()

            # Test snappy compression
            output_snappy = temp_dir / 'output_snappy.parquet'
            result = converter.convert(excel_file, output_snappy, compression='snappy')
            assert result is True
            assert output_snappy.exists()

            # Test no compression
            output_none = temp_dir / 'output_none.parquet'
            result = converter.convert(excel_file, output_none, compression='none')
            assert result is True
            assert output_none.exists()

    def test_convert_formulas_excel(self, converter, test_data_dir, temp_dir):
        """Test converting Excel with formulas."""
        excel_file = test_data_dir / 'formulas.xlsx'
        output_file = temp_dir / 'output.parquet'

        if excel_file.exists():
            # Should convert formula results, not formulas themselves
            result = converter.convert(excel_file, output_file)
            assert result is True
            assert output_file.exists()

    def test_convert_styled_excel(self, converter, test_data_dir, temp_dir):
        """Test converting styled Excel (ignores styling)."""
        excel_file = test_data_dir / 'styled.xlsx'
        output_file = temp_dir / 'output.parquet'

        if excel_file.exists():
            # Should extract data, ignoring styles
            result = converter.convert(excel_file, output_file)
            assert result is True
            assert output_file.exists()

    def test_convert_empty_sheets_handling(self, converter, test_data_dir, temp_dir):
        """Test handling of empty sheets."""
        excel_file = test_data_dir / 'empty_sheets.xlsx'
        output_dir = temp_dir / 'output'

        if excel_file.exists():
            result = converter.convert(excel_file, output_dir, separate=True)
            assert result is True

            # Check that non-empty sheets were converted
            assert (output_dir / 'empty_sheets_Data.parquet').exists()
            # Empty sheets might be skipped or create empty files

    def test_convert_duplicate_columns(self, converter, test_data_dir, temp_dir):
        """Test handling duplicate column names."""
        excel_file = test_data_dir / 'duplicate_columns.xlsx'
        output_file = temp_dir / 'output.parquet'

        if excel_file.exists():
            result = converter.convert(excel_file, output_file)
            # Duplicate columns cause Parquet schema errors - expected failure
            assert result is False
        else:
            pytest.skip("Test data file not found")

    def test_convert_error_handling(self, converter, temp_dir):
        """Test error handling during conversion."""
        # Create a corrupted Excel file
        corrupted_file = temp_dir / 'corrupted.xlsx'
        corrupted_file.write_bytes(b'Not a valid Excel file')

        output_file = temp_dir / 'output.parquet'
        result = converter.convert(corrupted_file, output_file)
        assert result is False

    def test_get_default_output_extension(self, converter):
        """Test default output extension."""
        # Check if converter has a default output extension method or use base class method
        ext = converter.get_output_extension('parquet')
        assert ext == '.parquet'

    def test_convert_with_sheet_names_special_chars(self, converter, test_data_dir, temp_dir):
        """Test handling sheet names with special characters."""
        excel_file = test_data_dir / 'multi_sheet.xlsx'
        output_dir = temp_dir / 'output'

        if excel_file.exists():
            result = converter.convert(excel_file, output_dir, separate=True)
            assert result is True

            # Check that files were created with sanitized names
            files = list(output_dir.glob('*.parquet'))
            assert len(files) > 0
            # All files should be valid paths
            for f in files:
                assert f.exists()

    def test_convert_output_directory_creation(self, converter, test_data_dir, temp_dir):
        """Test that output directory is created if it doesn't exist."""
        excel_file = test_data_dir / 'basic.xlsx'
        output_file = temp_dir / 'nested' / 'dir' / 'output.parquet'

        if excel_file.exists():
            # Directory doesn't exist yet
            assert not output_file.parent.exists()

            result = converter.convert(excel_file, output_file)
            assert result is True
            assert output_file.exists()
            assert output_file.parent.exists()

    def test_get_metadata_basic_excel(self, converter, test_data_dir):
        """Test metadata extraction from basic Excel file."""
        excel_file = test_data_dir / 'basic.xlsx'
        if excel_file.exists():
            metadata = converter.get_metadata(excel_file)

            # Check basic file information
            assert metadata is not None
            assert metadata['file_name'] == 'basic.xlsx'
            assert metadata['file_format'] == 'Excel Spreadsheet'
            assert metadata['file_extension'] == '.xlsx'
            assert metadata['file_size'] > 0
            assert 'modified_date' in metadata
            assert 'created_date' in metadata

            # Check sheet information
            assert 'sheet_count' in metadata
            assert 'sheet_names' in metadata
            assert metadata['sheet_count'] > 0
            assert isinstance(metadata['sheet_names'], list)

            # Check estimated data
            assert 'sheet_details' in metadata
            assert 'total_estimated_rows' in metadata
            assert 'max_columns' in metadata
        else:
            pytest.skip("Test data file not found")

    def test_get_metadata_multi_sheet_excel(self, converter, test_data_dir):
        """Test metadata extraction from multi-sheet Excel file."""
        excel_file = test_data_dir / 'multi_sheet.xlsx'
        if excel_file.exists():
            metadata = converter.get_metadata(excel_file)

            assert metadata is not None
            assert metadata['sheet_count'] > 1
            assert len(metadata['sheet_names']) > 1

            # Check sheet details
            sheet_details = metadata['sheet_details']
            assert len(sheet_details) > 1

            for _sheet_name, details in sheet_details.items():
                assert 'estimated_rows' in details
                assert 'estimated_columns' in details
                assert 'has_data' in details
                assert isinstance(details['has_data'], bool)
        else:
            pytest.skip("Test data file not found")

    def test_get_metadata_empty_excel(self, converter, temp_dir):
        """Test metadata extraction from empty Excel file."""
        from openpyxl import Workbook

        # Create an empty Excel file
        empty_file = temp_dir / 'empty.xlsx'
        wb = Workbook()
        wb.save(empty_file)
        wb.close()

        metadata = converter.get_metadata(empty_file)

        assert metadata is not None
        assert metadata['file_name'] == 'empty.xlsx'
        assert metadata['sheet_count'] == 1  # Default sheet
        assert metadata['sheet_names'] == ['Sheet']
        assert metadata['total_estimated_rows'] == 0
        assert metadata['max_columns'] == 0

        # Check that all sheets have no data
        for _sheet_name, details in metadata['sheet_details'].items():
            assert details['has_data'] is False
            assert details['estimated_rows'] == 0
            assert details['estimated_columns'] == 0

    def test_get_metadata_with_properties(self, converter, temp_dir):
        """Test metadata extraction with document properties."""
        from datetime import datetime

        from openpyxl import Workbook

        # Create an Excel file with properties
        excel_file = temp_dir / 'with_properties.xlsx'
        wb = Workbook()
        ws = wb.active

        # Add some data
        ws['A1'] = 'Name'
        ws['B1'] = 'Age'
        ws['A2'] = 'John'
        ws['B2'] = 25

        # Set document properties
        wb.properties.title = 'Test Document'
        wb.properties.creator = 'Test Author'
        wb.properties.subject = 'Test Subject'
        wb.properties.description = 'Test Description'
        wb.properties.keywords = 'test, metadata'
        wb.properties.category = 'Testing'
        # Note: openpyxl doesn't support company property in DocumentProperties
        wb.properties.created = datetime(2023, 1, 1, 12, 0, 0)
        wb.properties.modified = datetime(2023, 6, 1, 14, 30, 0)
        wb.properties.lastModifiedBy = 'Test Modifier'

        wb.save(excel_file)
        wb.close()

        metadata = converter.get_metadata(excel_file)

        assert metadata is not None
        assert metadata['title'] == 'Test Document'
        assert metadata['author'] == 'Test Author'
        assert metadata['subject'] == 'Test Subject'
        assert metadata['description'] == 'Test Description'
        assert metadata['keywords'] == 'test, metadata'
        assert metadata['category'] == 'Testing'
        assert metadata['company'] is None  # openpyxl doesn't support company property
        assert metadata['created'] == '2023-01-01T12:00:00'
        # Note: modified date gets updated when file is saved, so we just check it's present
        assert metadata['modified'] is not None
        assert len(metadata['modified']) > 10  # Should be ISO format string
        assert metadata['last_modified_by'] == 'Test Modifier'

        # Check data was counted correctly
        assert metadata['total_estimated_rows'] == 2
        assert metadata['max_columns'] == 2

    def test_get_metadata_corrupted_file(self, converter, temp_dir):
        """Test metadata extraction from corrupted Excel file."""
        # Create a corrupted Excel file
        corrupted_file = temp_dir / 'corrupted.xlsx'
        corrupted_file.write_bytes(b'Not a valid Excel file')

        metadata = converter.get_metadata(corrupted_file)

        # Should return None for corrupted files
        assert metadata is None

    def test_get_metadata_nonexistent_file(self, converter):
        """Test metadata extraction from non-existent file."""
        from pathlib import Path

        nonexistent_file = Path('/nonexistent/file.xlsx')
        metadata = converter.get_metadata(nonexistent_file)

        # Should return None for non-existent files
        assert metadata is None

    @patch('pyforge_cli.converters.excel_converter.HAS_OPENPYXL', False)
    def test_get_metadata_without_openpyxl(self, converter, temp_dir):
        """Test metadata extraction when openpyxl is not available."""
        from openpyxl import Workbook

        # Create a test Excel file
        excel_file = temp_dir / 'test.xlsx'
        wb = Workbook()
        wb.save(excel_file)
        wb.close()

        metadata = converter.get_metadata(excel_file)

        # Should return None when openpyxl is not available
        assert metadata is None

    def test_get_metadata_large_file_estimation(self, converter, temp_dir):
        """Test metadata extraction with large file size estimation."""
        from openpyxl import Workbook

        # Create an Excel file with more than 1000 rows
        excel_file = temp_dir / 'large.xlsx'
        wb = Workbook()
        ws = wb.active

        # Add headers
        ws['A1'] = 'Column1'
        ws['B1'] = 'Column2'
        ws['C1'] = 'Column3'

        # Add data beyond the 1000 row sampling limit
        for i in range(2, 1500):  # Add 1498 rows of data
            ws[f'A{i}'] = f'Row{i}'
            ws[f'B{i}'] = i * 2
            ws[f'C{i}'] = f'Data{i}'

        wb.save(excel_file)
        wb.close()

        metadata = converter.get_metadata(excel_file)

        assert metadata is not None
        assert metadata['sheet_count'] == 1

        # Should sample only first 1000 rows for estimation
        sheet_details = metadata['sheet_details']['Sheet']
        assert sheet_details['estimated_rows'] == 1000  # Limited by sampling
        assert sheet_details['estimated_columns'] == 3
        assert sheet_details['has_data'] is True

    def test_get_metadata_mixed_data_types(self, converter, temp_dir):
        """Test metadata extraction with mixed data types."""
        from datetime import datetime

        from openpyxl import Workbook

        # Create an Excel file with various data types
        excel_file = temp_dir / 'mixed_types.xlsx'
        wb = Workbook()
        ws = wb.active

        # Add headers
        ws['A1'] = 'Text'
        ws['B1'] = 'Number'
        ws['C1'] = 'Date'
        ws['D1'] = 'Boolean'
        ws['E1'] = 'Formula'

        # Add data with different types
        ws['A2'] = 'Sample text'
        ws['B2'] = 123.45
        ws['C2'] = datetime(2023, 1, 15)
        ws['D2'] = True
        ws['E2'] = '=B2*2'

        wb.save(excel_file)
        wb.close()

        metadata = converter.get_metadata(excel_file)

        assert metadata is not None
        assert metadata['max_columns'] == 5
        assert metadata['total_estimated_rows'] == 2

        sheet_details = metadata['sheet_details']['Sheet']
        assert sheet_details['estimated_columns'] == 5
        assert sheet_details['has_data'] is True
