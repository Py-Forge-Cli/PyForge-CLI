#!/usr/bin/env python
"""
Generate Excel test files with various edge cases for comprehensive testing.
"""

import datetime
import random
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


def generate_basic_excel():
    """Generate a basic Excel file with simple data."""
    df = pd.DataFrame(
        {
            "Name": ["Alice", "Bob", "Charlie", "David"],
            "Age": [25, 30, 35, 40],
            "City": ["New York", "Los Angeles", "Chicago", "Houston"],
            "Salary": [50000, 60000, 70000, 80000],
        }
    )
    return df


def generate_multi_sheet_excel():
    """Generate Excel with multiple sheets."""
    sheets = {
        "Employees": pd.DataFrame(
            {
                "ID": range(1, 11),
                "Name": [f"Employee_{i}" for i in range(1, 11)],
                "Department": [
                    "Sales",
                    "IT",
                    "HR",
                    "Sales",
                    "IT",
                    "Finance",
                    "HR",
                    "Sales",
                    "IT",
                    "Finance",
                ],
                "Salary": [random.randint(40000, 100000) for _ in range(10)],
            }
        ),
        "Departments": pd.DataFrame(
            {
                "DeptID": [1, 2, 3, 4],
                "DeptName": ["Sales", "IT", "HR", "Finance"],
                "Manager": ["John Doe", "Jane Smith", "Bob Johnson", "Alice Brown"],
            }
        ),
        "Projects": pd.DataFrame(
            {
                "ProjectID": range(1, 6),
                "ProjectName": [f"Project_{chr(65+i)}" for i in range(5)],
                "Budget": [random.randint(10000, 100000) for _ in range(5)],
                "StartDate": pd.date_range("2023-01-01", periods=5, freq="M"),
            }
        ),
    }
    return sheets


def generate_empty_sheets_excel():
    """Generate Excel with some empty sheets."""
    sheets = {
        "Data": pd.DataFrame({"A": [1, 2, 3], "B": [4, 5, 6]}),
        "EmptySheet": pd.DataFrame(),
        "SingleColumn": pd.DataFrame({"Col1": [1, 2, 3]}),
        "SingleRow": pd.DataFrame({"A": [1], "B": [2], "C": [3]}),
        "AnotherEmpty": pd.DataFrame(),
    }
    return sheets


def generate_large_excel():
    """Generate Excel with large dataset."""
    # Create a large dataset (10000 rows)
    data = {
        "ID": range(1, 10001),
        "Timestamp": pd.date_range("2020-01-01", periods=10000, freq="h"),
        "Value": np.random.randn(10000),
        "Category": [random.choice(["A", "B", "C", "D", "E"]) for _ in range(10000)],
        "Status": [
            random.choice(["Active", "Inactive", "Pending"]) for _ in range(10000)
        ],
    }
    return pd.DataFrame(data)


def generate_mixed_types_excel():
    """Generate Excel with mixed data types including edge cases."""
    df = pd.DataFrame(
        {
            "Integer": [1, 2, 3, 4, 5],
            "Float": [1.1, 2.2, 3.3, 4.4, 5.5],
            "String": ["Hello", "World", "Test", "Data", "Excel"],
            "Date": pd.date_range("2023-01-01", periods=5),
            "DateTime": pd.date_range("2023-01-01 10:00:00", periods=5, freq="h"),
            "Boolean": [True, False, True, False, True],
            "Mixed": [1, "two", 3.0, True, None],
            "Currency": ["$100.00", "$200.50", "$300.75", "$400.25", "$500.00"],
            "Percentage": ["10%", "20%", "30%", "40%", "50%"],
        }
    )
    return df


def generate_special_characters_excel():
    """Generate Excel with special characters and unicode."""
    df = pd.DataFrame(
        {
            "Normal": ["Test1", "Test2", "Test3"],
            "Unicode": ["Hello 世界", "Привет мир", "مرحبا بالعالم"],
            "Special": ["Test@#$", "Data&*()%", "Value!^~"],
            "Newlines": ["Line1\nLine2", "First\nSecond\nThird", "Single"],
            "Tabs": ["Tab\tSeparated", "Multiple\t\tTabs", "Normal"],
            "Quotes": ['"Double"', "'Single'", """Both"'quotes"""],
        }
    )
    return df


def generate_formulas_excel():
    """Generate Excel with formulas (requires special handling)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Formulas"

    # Add headers
    headers = ["A", "B", "Sum", "Average", "Max", "Concatenated"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Add data and formulas
    for row in range(2, 12):
        ws.cell(row=row, column=1, value=random.randint(1, 100))
        ws.cell(row=row, column=2, value=random.randint(1, 100))
        ws.cell(row=row, column=3, value=f"=A{row}+B{row}")
        ws.cell(row=row, column=4, value=f"=AVERAGE(A{row}:B{row})")
        ws.cell(row=row, column=5, value=f"=MAX(A{row}:B{row})")
        ws.cell(row=row, column=6, value=f'=CONCATENATE(A{row},"-",B{row})')

    return wb


def generate_styled_excel():
    """Generate Excel with formatting and styles."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Styled"

    # Add data
    data = [
        ["Header 1", "Header 2", "Header 3"],
        ["Data 1", 100, datetime.date.today()],
        ["Data 2", 200, datetime.date.today() - datetime.timedelta(days=1)],
        ["Data 3", 300, datetime.date.today() - datetime.timedelta(days=2)],
    ]

    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Style headers
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")

            # Style data cells
            elif col_idx == 2:
                cell.number_format = "#,##0.00"
            elif col_idx == 3:
                cell.number_format = "yyyy-mm-dd"

    # Add a chart
    chart = BarChart()
    chart.title = "Sample Chart"
    chart.y_axis.title = "Values"
    chart.x_axis.title = "Categories"

    data = Reference(ws, min_col=2, min_row=1, max_row=4, max_col=2)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E5")

    return wb


def generate_pivot_like_excel():
    """Generate Excel that simulates pivot table structure."""
    # Create data that looks like a pivot table
    data = []
    regions = ["North", "South", "East", "West"]
    products = ["Product A", "Product B", "Product C"]
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]

    for region in regions:
        for product in products:
            row = {"Region": region, "Product": product}
            for month in months:
                row[month] = random.randint(100, 1000)
            row["Total"] = sum(row[m] for m in months)
            data.append(row)

    df = pd.DataFrame(data)
    return df


def generate_null_heavy_excel():
    """Generate Excel with many null/empty values."""
    df = pd.DataFrame(
        {
            "Complete": range(1, 101),
            "Sparse1": [i if i % 5 == 0 else None for i in range(1, 101)],
            "Sparse2": [i if i % 10 == 0 else np.nan for i in range(1, 101)],
            "MostlyEmpty": [i if i > 95 else None for i in range(1, 101)],
            "RandomNulls": [
                i if random.random() > 0.7 else None for i in range(1, 101)
            ],
        }
    )
    return df


def generate_duplicate_columns_excel():
    """Generate Excel with duplicate column names."""
    # Create DataFrame with duplicate column names
    df = pd.DataFrame(
        {
            "ID": range(1, 6),
            "Name": ["A", "B", "C", "D", "E"],
            "Value": [10, 20, 30, 40, 50],
        }
    )
    # Add duplicate columns
    df2 = pd.DataFrame(
        {"Name": ["X", "Y", "Z", "W", "V"], "Value": [100, 200, 300, 400, 500]}
    )
    # Combine with duplicate names
    result = pd.concat([df, df2], axis=1)
    return result


def generate_hierarchical_excel():
    """Generate Excel with hierarchical/nested structure."""
    # Create multi-level column structure
    arrays = [
        ["Sales", "Sales", "Sales", "HR", "HR", "IT", "IT"],
        ["Q1", "Q2", "Q3", "Hired", "Left", "Projects", "Tickets"],
    ]
    tuples = list(zip(*arrays))
    index = pd.MultiIndex.from_tuples(tuples)

    df = pd.DataFrame(
        np.random.randint(0, 100, size=(10, 7)),
        columns=index,
        index=[f"Row_{i}" for i in range(1, 11)],
    )
    return df


def save_test_files():
    """Save all test Excel files."""
    output_dir = Path(__file__).parent.parent / "data" / "excel"
    output_dir.mkdir(parents=True, exist_ok=True)

    # Simple DataFrames
    test_cases = [
        ("basic.xlsx", generate_basic_excel()),
        ("large_dataset.xlsx", generate_large_excel()),
        ("mixed_types.xlsx", generate_mixed_types_excel()),
        ("special_characters.xlsx", generate_special_characters_excel()),
        ("pivot_like.xlsx", generate_pivot_like_excel()),
        ("null_heavy.xlsx", generate_null_heavy_excel()),
        ("duplicate_columns.xlsx", generate_duplicate_columns_excel()),
        # ('hierarchical.xlsx', generate_hierarchical_excel())  # Skip due to MultiIndex issue
    ]

    for filename, df in test_cases:
        filepath = output_dir / filename
        if isinstance(df, pd.DataFrame):
            df.to_excel(filepath, index=False)
        print(f"Created: {filepath}")

    # Multi-sheet files
    multi_sheet_cases = [
        ("multi_sheet.xlsx", generate_multi_sheet_excel()),
        ("empty_sheets.xlsx", generate_empty_sheets_excel()),
    ]

    for filename, sheets_dict in multi_sheet_cases:
        filepath = output_dir / filename
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            for sheet_name, df in sheets_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"Created: {filepath}")

    # Special workbook files (with formulas/styles)
    special_cases = [
        ("formulas.xlsx", generate_formulas_excel()),
        ("styled.xlsx", generate_styled_excel()),
    ]

    for filename, wb in special_cases:
        filepath = output_dir / filename
        wb.save(filepath)
        print(f"Created: {filepath}")

    print(
        f"\nGenerated {len(test_cases) + len(multi_sheet_cases) + len(special_cases)} Excel test files in {output_dir}"
    )


if __name__ == "__main__":
    save_test_files()
